import numpy as np
import random
import pandas as pd
import datetime
import openpyxl


time_slots = ['08:00-10:00', '10:00-12:00', '12:00-14:00', '14:00-16:00', '16:00-18:00']
num_days = 10 # the number of exam days
slots_per_day = len(time_slots)
total_slots = num_days * slots_per_day

start_date = datetime.date.today()  # yyyy, mm, dd
day_buffer = 14
# Generate a list of dates using a loop and the datetime module
exam_days = []
for i in range(num_days):
    day = start_date + datetime.timedelta(days=i+14)
    exam_days.append(day.strftime('%Y-%m-%d'))

generated_available_timeslots = []
for day in exam_days:
    for slot in time_slots:
        generated_available_timeslots.append(day + ' ' + slot)

# read student_class.xlsx
df_student_class = pd.read_excel('Data/student_class.xlsx')

# read classes.xlsx
df_classes = pd.read_excel('Data/classes.xlsx')

# read classrooms.xlsx
df_classrooms = pd.read_excel('Data/classrooms.xlsx')


# create rooms dictionary
rooms = []
for i, row in df_classrooms.iterrows():
    room_dict = {'id': row['DERSLİK KODU'], 'capacity': row['SINAV KAPASİTESİ'], 'availability': [], 'schedule': {}}
    # Add room details
    room_dict['building'] = row['BİNA ADI']
    room_dict['floor'] = row['KAT']
    room_dict['room_code'] = row['DERSLİK KODU']
    
    # Initialize the schedule for each day and timeslot
    for day in exam_days:
        schedule = {}
        for time_slot in time_slots:
            schedule[time_slot] = {'course': None, 'instructor': None}
        room_dict['schedule'][day] = schedule
    rooms.append(room_dict)

rooms_capacities = []
for room in rooms:
    rooms_capacities.append(room['capacity'])

print(max(rooms_capacities))


# create exams dictionary
exams = []
for exam in df_classes['Ders Kodu'].unique():
    num_students = len(df_student_class[df_student_class['Ders Kodu'] == exam])
    exam_dict = {'id': exam, 'num_students': num_students}
    
    # Find a room with enough capacity to accommodate the exam
    for room in rooms:
        if room['capacity'] >= num_students:
            # Assign the exam to the room
            room['schedule'][exam_days[0]][time_slots[0]]['course'] = exam + '_room'
            exam_dict['assigned_room'] = room['id']
            break
    
    exams.append(exam_dict)

print(len(exams))


# create students dictionary
students = []
for student in df_student_class['Öğrenci No'].unique():
    taken_exams = list(df_student_class[df_student_class['Öğrenci No'] == student]['Ders Kodu'].unique())
    student_dict = {'id': str(student), 'taken_exams': taken_exams}
    students.append(student_dict)


# Define the constraint (non-conforming exam)
non_conforming_exam = {'BAU091', 'BAU107', 'BWL407', 'BWL413', 'BWL415', 'BWL417', 'ENG203', 'DEU121', 'WIN311', 'WIN313' }


def calculate_cost(solution):
    # Initialize the cost to zero
    cost = 0
    # Initialize a set to keep track of students' exam timeslots
    student_timeslots = set()

    # Count the number of students taking exams at the same time in the same room
    for timeslot in generated_available_timeslots:
        day, time = timeslot.split()
        for room in rooms:
            students_in_room = []
            exams_in_room = room['schedule'][day][time]['course']
            for exam, assigned_timeslot in solution.items():
                if not exam.endswith("_room"):
                        exam_day, exam_time = assigned_timeslot.split(" ")
                        if exam_day == day and exam_time == time:
                            if exam in non_conforming_exam:
                                cost += 1
                            if exam == exams_in_room:
                                students_in_room.extend([student['id'] for student in students if exam in student['taken_exams']])

                                #Add room details to the cost calculation
                                room_id = room['id']
                                building = room['building']
                                floor = room['floor']
                                cost += 0.1 * len(students_in_room) * (len(students_in_room) - 1) / 2 

                            else:
                                cost += 1

                            # Check if any student is already assigned to the current timeslot
                            for student_id in students_in_room:
                                if (student_id, day, time) in student_timeslots:
                                    cost += 1  # Increase the cost if a student has multiple exams at the same timeslot
                                else:
                                    student_timeslots.add((student_id, day, time))


                cost += len(students_in_room) * (len(students_in_room) - 1) / 2
    return cost



# Define the perturbation function
def perturb_solution(solution):
    # Create a copy of the solution to modify
    new_solution = solution.copy()
    # Choose an exam to reassign to a new timeslot
    exam_id = random.choice(list(key for key in new_solution.keys() if not key.endswith("_room")))
    new_timeslot = random.choice(generated_available_timeslots)
    # Assign the exam to the new timeslot
    new_solution[exam_id] = new_timeslot

    room = random.choice(rooms)
    room['schedule'][day][time]['course'] = exam_id
    # Assign the exam to the new timeslot
    new_solution[exam_id + '_room'] = {'id': room['id'], 'building': room['building'], 'floor': room['floor'], 'room_code': room['room_code']}

    return new_solution


def simulated_annealing(current_solution, initial_temperature, cooling_factor, num_iterations):
    # Set the initial temperature
    current_temperature = initial_temperature
    # Set the current solution and cost as the best solution and cost found so far
    best_solution = current_solution
    best_cost = calculate_cost(current_solution)
    # Calculate the cost of the initial solution
    current_cost = calculate_cost(current_solution)
    # Iterate for the specified number of iterations
    for i in range(num_iterations):
        # Perturb the current solution
        new_solution = perturb_solution(current_solution)
        # Calculate the cost of the new solution
        new_cost = calculate_cost(new_solution)
        # Calculate the change in cost
        delta_cost = new_cost - current_cost
        # If the new solution is better, accept it
        if delta_cost < 0:
            current_solution = new_solution
            current_cost = new_cost
            # If the new solution is the best found so far, update the best solution
            if current_cost < best_cost:
                best_solution = current_solution
                best_cost = current_cost
        # If the new solution is worse, accept it with a certain probability
        else:
            acceptance_probability = np.exp(-delta_cost / current_temperature)
            if random.random() < acceptance_probability:
                current_solution = new_solution
                current_cost = new_cost
        # Reduce the temperature
        current_temperature *= cooling_factor
    # Return the best solution found
    return best_solution, best_cost


exams_dict = {exam['id']: exam for exam in exams}
rooms_dict = {room['id']: room for room in rooms}
students_dict = {student['id']: student for student in students}

# 197 exams, 77 classrooms, 2130 students

# create an initial solution by assigning exams to the available timeslots
initial_solution = {}
for exam in exams:
    assigned = False
    while not assigned:
        # Choose a random room and timeslot
        room = random.choice(rooms)
        day, time = random.choice(generated_available_timeslots).split(" ")
        # Check if the room is available
        if exam['num_students'] > max(rooms_capacities):
            print(exam['id'] + '  This class is taken by students more than max capacity of one biggest classroom')
            assigned = True
        if room['schedule'][day][time]['course'] is None and exam['num_students'] <= room['capacity'] :
            # Assign the exam to the room and timeslot
            room['schedule'][day][time]['course'] = exam['id']
            initial_solution[exam['id']] = day + ' ' + time

            # Add room details to the initial solution
            initial_solution[exam['id']+'_room'] = {'id': room['id'], 'building': room['building'], 'floor': room['floor'], 'room_code': room['room_code']}
            
            assigned = True

#print(initial_solution)

optimized_solution, best_cost = simulated_annealing(initial_solution, initial_temperature=10, cooling_factor=0.99, num_iterations=2000)

# print("Optimized solution:")
# best-cost: 2991058.8/ initial_temp = 100/ cooling_factor=0.95 / num_iter=1000 ---> 15396.0
print(optimized_solution, best_cost)


"""
def find_dict_differences(dict1, dict2):
    return {k: (v, dict2[k]) for k, v in dict1.items() if v != dict2[k]}

print(find_dict_differences(initial_solution, optimized_solution))
"""

# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Write column headers
sheet['A1'] = 'Exam ID'
sheet['B1'] = 'Timeslot'
sheet['C1'] = 'Room Info'

# Write dictionary values to rows
row = 2  # start writing rows from the second row
for key, value in initial_solution.items():
    if not key.endswith("_room"):
        # Write exam ID to column A
        sheet.cell(row=row, column=1).value = key
        
        # Write timeslot to column B
        timeslot = value
        if type(value) == dict:
            timeslot = value.get('timeslot')
        sheet.cell(row=row, column=2).value = timeslot
    if key.endswith("_room"):
        # Write room info to column C
        room_info = ''
        if type(value) == dict:
            room_info = f"{value.get('room_code')}, {value.get('building')}, Floor {value.get('floor')}"
        sheet.cell(row=row, column=3).value = room_info

        row += 1
    
    

# Save the Excel workbook
wb.save('initial_exam_schedule.xlsx')

# Create a new Excel workbook and select the active sheet
wb2 = openpyxl.Workbook()
sheet = wb2.active

# Write column headers
sheet['A1'] = 'Exam ID'
sheet['B1'] = 'Timeslot'
sheet['C1'] = 'Room Info'

# Write dictionary values to rows
row = 2  # start writing rows from the second row
for key, value in optimized_solution.items():
    if not key.endswith("_room"):
        # Write exam ID to column A
        sheet.cell(row=row, column=1).value = key
        
        # Write timeslot to column B
        timeslot = value
        if type(value) == dict:
            timeslot = value.get('timeslot')
        sheet.cell(row=row, column=2).value = timeslot
    if key.endswith("_room"):
        # Write room info to column C
        room_info = ''
        if type(value) == dict:
            room_info = f"{value.get('room_code')}, {value.get('building')}, Floor {value.get('floor')}"
        sheet.cell(row=row, column=3).value = room_info
    
        row += 1
    


# Save the Excel workbook
wb2.save('exam_schedule.xlsx')

"""
#INF205,211,WIN107
for student in students:
    for exam in student['taken_exams']:
        if exam not in ['DEU121', 'ENG101', 'ENG201', 'ENG301', 'MAT103', 'MAT201']:
            print(exam)
            print(optimized_solution[exam])
            print(optimized_solution[str(exam) + "_room"])
    print("--------")
"""